"""
Platform Logs Enterprise Excel Report Generator
================================================
4-sheet workbook:
  Sheet 1  "Activity Log"        — all records, severity-coded rows
  Sheet 2  "System Intelligence" — algorithmic analysis with visual bars
  Sheet 3  "Alerts & Anomalies"  — warning / error / critical only
  Sheet 4  "Company Report"      — per-company summary table
"""

import io
import json
from collections import Counter, defaultdict
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Protection, Side,
)
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────
_BRAND_BLUE  = "1E3A8A"
_BRAND_DARK  = "0F172A"
_BRAND_SLATE = "334155"

_FILL_TITLE   = PatternFill("solid", fgColor=_BRAND_BLUE)
_FILL_COLHDR  = PatternFill("solid", fgColor="1E3A5F")
_FILL_SECTION = PatternFill("solid", fgColor="EFF6FF")
_FILL_ZEBRA   = PatternFill("solid", fgColor="F8FAFC")
_FILL_WHITE   = PatternFill("solid", fgColor="FFFFFF")
_FILL_BLANK   = PatternFill("solid", fgColor="F1F5F9")

# Severity: (bg, fg, bar-color, border-accent)
_SEV = {
    "info":     ("DBEAFE", "1E40AF", "3B82F6", "BFDBFE"),
    "warning":  ("FFFBEB", "92400E", "F59E0B", "FDE68A"),
    "error":    ("FEF2F2", "991B1B", "EF4444", "FECACA"),
    "critical": ("FDF2F8", "9D174D", "EC4899", "FBCFE8"),
}

# Category: (bg, fg)
_CAT = {
    "system":    ("F1F5F9", "475569"),
    "import":    ("EDE9FE", "5B21B6"),
    "warehouse": ("CCFBF1", "0F766E"),
    "edit":      ("F3E8FF", "6D28D9"),
    "alert":     ("FFEDD5", "C2410C"),
    "analysis":  ("CFFAFE", "0E7490"),
}

# Fonts
_F_TITLE    = Font(name="Calibri", bold=True,  color="FFFFFF",   size=16)
_F_SUBTITLE = Font(name="Calibri",              color="93C5FD",   size=10)
_F_COLHDR   = Font(name="Calibri", bold=True,  color="E0E7FF",   size=11)
_F_DATA     = Font(name="Calibri",              color="1E293B",   size=10)
_F_BOLD     = Font(name="Calibri", bold=True,  color="1E293B",   size=10)
_F_MUTED    = Font(name="Calibri",              color="64748B",   size=9)
_F_SECTION  = Font(name="Calibri", bold=True,  color=_BRAND_BLUE, size=12)
_F_METRIC   = Font(name="Calibri", bold=True,  color=_BRAND_DARK, size=20)
_F_MLABEL   = Font(name="Calibri",              color="64748B",   size=8)
_F_MONO     = Font(name="Courier New",          color="334155",   size=9)
_F_SAFE     = Font(name="Calibri", italic=True, color="166534",   size=9)

# Shared
_THIN   = Border(
    left=Side("thin",  color="E2E8F0"), right=Side("thin",  color="E2E8F0"),
    top=Side("thin",   color="E2E8F0"), bottom=Side("thin", color="E2E8F0"),
)
_WRAP   = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", indent=1)
_LOCK   = Protection(locked=True)


# ── Helpers ──────────────────────────────────────────────────────────────

def _lock(ws):
    ws.protection.sheet = True
    ws.protection.password = "safeware"
    ws.protection.sort = False
    ws.protection.autoFilter = False


def _header_block(ws, title: str, sub: str, nc: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    c = ws.cell(1, 1, title)
    c.font = _F_TITLE; c.fill = _FILL_TITLE; c.alignment = _CENTER; c.protection = _LOCK
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    c2 = ws.cell(2, 1, sub)
    c2.font = _F_SUBTITLE; c2.fill = _FILL_TITLE; c2.alignment = _CENTER; c2.protection = _LOCK
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 5
    for ci in range(1, nc + 1):
        ws.cell(3, ci).fill = _FILL_TITLE; ws.cell(3, ci).protection = _LOCK


def _col_headers(ws, row: int, cols: list):
    for ci, (text, width) in enumerate(cols, 1):
        c = ws.cell(row, ci, text)
        c.font = _F_COLHDR; c.fill = _FILL_COLHDR
        c.alignment = _CENTER; c.border = _THIN; c.protection = _LOCK
        ws.column_dimensions[get_column_letter(ci)].width = width


def _section_row(ws, row: int, text: str, nc: int, bg="EFF6FF", fg=_BRAND_BLUE, ht=24):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row, 1, text)
    c.font = Font(name="Calibri", bold=True, size=12, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = _LEFT; c.border = _THIN; c.protection = _LOCK
    ws.row_dimensions[row].height = ht


def _blank_row(ws, row: int, nc: int):
    for ci in range(1, nc + 1):
        ws.cell(row, ci).fill = _FILL_BLANK; ws.cell(row, ci).protection = _LOCK
    ws.row_dimensions[row].height = 8


def _bar(count: int, max_count: int, width: int = 44) -> str:
    if max_count == 0 or count == 0:
        return "░" * width
    filled = round(count / max_count * width)
    return "█" * filled + "░" * (width - filled)


def _pct(part: int, total: int) -> str:
    if total == 0:
        return "0.0%"
    return f"{part / total * 100:.1f}%"


def _fmt_ts(ts: Optional[str]) -> str:
    if not ts:
        return ""
    return ts.replace("T", " ").rstrip("Z")


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 1 — Activity Log
# ═══════════════════════════════════════════════════════════════════════

_S1_COLS = [
    ("#",          5),
    ("Timestamp",  20),
    ("Company",    20),
    ("User",       22),
    ("Category",   14),
    ("Severity",   12),
    ("Event Type", 22),
    ("Title",      38),
    ("Detail",     44),
    ("Entity",     20),
    ("IP Address", 16),
]


def _build_activity_sheet(ws, logs: list, ts: str, total: int):
    nc = len(_S1_COLS)
    _header_block(ws, "SAFEWARE — PLATFORM ACTIVITY LOG", f"Generated: {ts}  |  Records: {total:,}  |  CONFIDENTIAL — Internal Use Only", nc)
    hdr = 4
    _col_headers(ws, hdr, _S1_COLS)

    for idx, log in enumerate(logs):
        r = hdr + 1 + idx
        sev = (log.get("severity") or "info").lower()
        bg, fg, _, _ = _SEV.get(sev, _SEV["info"])
        row_fill = PatternFill("solid", fgColor=bg) if sev in ("error", "critical") else (
            _FILL_ZEBRA if idx % 2 else _FILL_WHITE
        )

        vals = [
            idx + 1,
            _fmt_ts(log.get("created_at")),
            log.get("company_name") or "",
            log.get("user_id") or "—",
            (log.get("category") or "system").upper(),
            sev.upper(),
            log.get("event_type") or "",
            log.get("title") or "",
            log.get("detail") or "",
            (f"{log.get('entity_type') or ''}: {log.get('entity_name') or ''}").strip(": ") or "—",
            log.get("ip_address") or "—",
        ]

        # Left accent border by severity
        left_color = _SEV.get(sev, _SEV["info"])[2]
        left_border = Border(
            left=Side("medium", color=left_color),
            right=Side("thin", color="E2E8F0"),
            top=Side("thin", color="F1F5F9"),
            bottom=Side("thin", color="F1F5F9"),
        )

        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.fill = row_fill
            c.protection = _LOCK
            c.border = left_border if ci == 1 else _THIN
            if ci in (2, 7, 11):  # timestamp, event_type, ip — monospace
                c.font = _F_MONO
                c.alignment = Alignment(vertical="center")
            elif ci == 8:  # title — bold
                c.font = Font(name="Calibri", size=10, color=fg if sev in ("error", "critical") else "1E293B", bold=(sev in ("error", "critical")))
                c.alignment = _WRAP
            else:
                c.font = _F_DATA
                c.alignment = _WRAP
        ws.row_dimensions[r].height = 18

    last = hdr + max(len(logs), 1)
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(nc)}{last}"
    ws.freeze_panes = f"A{hdr + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = f"1:{hdr}"
    _lock(ws)


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 2 — System Intelligence
# ═══════════════════════════════════════════════════════════════════════

def _build_intelligence_sheet(ws, logs: list, stats: dict, ts: str):
    NC = 4
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 52

    total = stats.get("total", len(logs))
    by_sev = stats.get("by_severity", {})
    by_cat = stats.get("by_category", {})
    recent = stats.get("recent_24h", 0)
    companies_count = stats.get("companies_count", 0)
    per_company = stats.get("per_company", {})
    errs = stats.get("errors_warnings", 0)
    err_rate = f"{errs / total * 100:.1f}%" if total else "0.0%"

    # Compute from logs
    event_counter = Counter(log.get("event_type") or "unknown" for log in logs)
    top_events = event_counter.most_common(10)
    user_counter = Counter(
        log.get("user_id") or "system"
        for log in logs
        if log.get("user_id") and log.get("user_id") != "system"
    )
    top_users = user_counter.most_common(5)

    _header_block(ws, "SYSTEM INTELLIGENCE REPORT", f"Generated: {ts}  |  Algorithmic analysis of {total:,} events  |  SAFEWARE Platform", NC)

    row = 4

    # ── Section A: Executive Summary ────────────────────────────────────
    _section_row(ws, row, "▌ A.  EXECUTIVE SUMMARY", NC, bg="EFF6FF"); row += 1

    metrics_left = [
        ("Total Events",      f"{total:,}"),
        ("Errors & Warnings", f"{errs:,}"),
        ("Error Rate",        err_rate),
    ]
    metrics_right = [
        ("Active Today (24h)", f"{recent:,}"),
        ("Companies Monitored", f"{companies_count}"),
        ("Report Generated",   ts[:10]),
    ]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 52

    for mi, ((lk, lv), (rk, rv)) in enumerate(zip(metrics_left, metrics_right)):
        r = row + mi
        # Label left
        cl = ws.cell(r, 1, lk)
        cl.font = _F_MLABEL; cl.fill = _FILL_WHITE; cl.border = _THIN; cl.protection = _LOCK
        cl.alignment = _LEFT
        # Value left
        cv = ws.cell(r, 2, lv)
        cv.font = Font(name="Calibri", bold=True, size=13, color=_BRAND_DARK)
        cv.fill = _FILL_WHITE; cv.border = _THIN; cv.protection = _LOCK; cv.alignment = _CENTER
        # Label right (use col 3)
        cr = ws.cell(r, 3, rk)
        cr.font = _F_MLABEL; cr.fill = _FILL_WHITE; cr.border = _THIN; cr.protection = _LOCK
        cr.alignment = _LEFT
        # Value right (use col 4)
        crv = ws.cell(r, 4, rv)
        crv.font = Font(name="Calibri", bold=True, size=13, color=_BRAND_DARK)
        crv.fill = _FILL_WHITE; crv.border = _THIN; crv.protection = _LOCK; crv.alignment = _CENTER
        ws.row_dimensions[r].height = 22
    row += 3

    _blank_row(ws, row, NC); row += 1

    # ── Section B: Severity Distribution ────────────────────────────────
    _section_row(ws, row, "▌ B.  SEVERITY DISTRIBUTION", NC); row += 1

    # header
    for ci, (txt, _) in enumerate([("Severity", 0), ("Count", 0), ("% Total", 0), ("Distribution", 0)], 1):
        c = ws.cell(row, ci, txt)
        c.font = _F_COLHDR; c.fill = _FILL_COLHDR; c.alignment = _CENTER
        c.border = _THIN; c.protection = _LOCK
    ws.row_dimensions[row].height = 20; row += 1

    sev_order = ["info", "warning", "error", "critical"]
    max_sev = max((by_sev.get(s, 0) for s in sev_order), default=1)
    for sev in sev_order:
        cnt = by_sev.get(sev, 0)
        bg, fg, bar_c, _ = _SEV[sev]

        ca = ws.cell(row, 1, sev.upper())
        ca.font = Font(name="Calibri", bold=True, size=10, color=fg)
        ca.fill = PatternFill("solid", fgColor=bg)
        ca.alignment = _LEFT; ca.border = _THIN; ca.protection = _LOCK

        cb = ws.cell(row, 2, f"{cnt:,}")
        cb.font = Font(name="Calibri", bold=True, size=11, color=fg)
        cb.fill = PatternFill("solid", fgColor=bg)
        cb.alignment = _CENTER; cb.border = _THIN; cb.protection = _LOCK

        cc = ws.cell(row, 3, _pct(cnt, total))
        cc.font = _F_MUTED; cc.fill = PatternFill("solid", fgColor=bg)
        cc.alignment = _CENTER; cc.border = _THIN; cc.protection = _LOCK

        cd = ws.cell(row, 4, _bar(cnt, max_sev))
        cd.font = Font(name="Courier New", size=10, color=bar_c)
        cd.fill = PatternFill("solid", fgColor=bg)
        cd.alignment = Alignment(horizontal="left", vertical="center")
        cd.border = _THIN; cd.protection = _LOCK
        ws.row_dimensions[row].height = 20; row += 1

    _blank_row(ws, row, NC); row += 1

    # ── Section C: Activity by Category ─────────────────────────────────
    _section_row(ws, row, "▌ C.  ACTIVITY BY CATEGORY", NC); row += 1

    for ci, txt in enumerate(["Category", "Count", "% Total", "Distribution"], 1):
        c = ws.cell(row, ci, txt)
        c.font = _F_COLHDR; c.fill = _FILL_COLHDR; c.alignment = _CENTER
        c.border = _THIN; c.protection = _LOCK
    ws.row_dimensions[row].height = 20; row += 1

    max_cat = max(by_cat.values(), default=1)
    cat_order = sorted(by_cat.keys(), key=lambda k: by_cat[k], reverse=True)
    for cat in cat_order:
        cnt = by_cat[cat]
        bg_cat, fg_cat = _CAT.get(cat, ("F1F5F9", "475569"))
        bar_color = fg_cat

        ca = ws.cell(row, 1, cat.upper())
        ca.font = Font(name="Calibri", bold=True, size=10, color=fg_cat)
        ca.fill = PatternFill("solid", fgColor=bg_cat)
        ca.alignment = _LEFT; ca.border = _THIN; ca.protection = _LOCK

        cb = ws.cell(row, 2, f"{cnt:,}")
        cb.font = Font(name="Calibri", bold=True, size=11, color=fg_cat)
        cb.fill = PatternFill("solid", fgColor=bg_cat)
        cb.alignment = _CENTER; cb.border = _THIN; cb.protection = _LOCK

        cc = ws.cell(row, 3, _pct(cnt, total))
        cc.font = _F_MUTED; cc.fill = PatternFill("solid", fgColor=bg_cat)
        cc.alignment = _CENTER; cc.border = _THIN; cc.protection = _LOCK

        cd = ws.cell(row, 4, _bar(cnt, max_cat))
        cd.font = Font(name="Courier New", size=10, color=bar_color)
        cd.fill = PatternFill("solid", fgColor=bg_cat)
        cd.alignment = Alignment(horizontal="left", vertical="center")
        cd.border = _THIN; cd.protection = _LOCK
        ws.row_dimensions[row].height = 20; row += 1

    _blank_row(ws, row, NC); row += 1

    # ── Section D: Top 10 Event Types ────────────────────────────────────
    _section_row(ws, row, "▌ D.  TOP EVENT TYPES (by frequency)", NC); row += 1

    for ci, txt in enumerate(["Event Type", "Count", "% Total", "Frequency Bar"], 1):
        c = ws.cell(row, ci, txt)
        c.font = _F_COLHDR; c.fill = _FILL_COLHDR; c.alignment = _CENTER
        c.border = _THIN; c.protection = _LOCK
    ws.row_dimensions[row].height = 20; row += 1

    max_evt = top_events[0][1] if top_events else 1
    for rank, (evt, cnt) in enumerate(top_events, 1):
        is_err = any(x in evt for x in ("fail", "error", "block", "suspend", "reject"))
        fg_evt = "991B1B" if is_err else "1E293B"
        fill_evt = _FILL_ZEBRA if rank % 2 else _FILL_WHITE

        ca = ws.cell(row, 1, evt)
        ca.font = Font(name="Courier New", size=10, color=fg_evt)
        ca.fill = fill_evt; ca.alignment = _LEFT; ca.border = _THIN; ca.protection = _LOCK

        cb = ws.cell(row, 2, f"{cnt:,}")
        cb.font = _F_BOLD; cb.fill = fill_evt; cb.alignment = _CENTER
        cb.border = _THIN; cb.protection = _LOCK

        cc = ws.cell(row, 3, _pct(cnt, total))
        cc.font = _F_MUTED; cc.fill = fill_evt; cc.alignment = _CENTER
        cc.border = _THIN; cc.protection = _LOCK

        cd = ws.cell(row, 4, _bar(cnt, max_evt, 40))
        cd.font = Font(name="Courier New", size=10, color="2563EB" if not is_err else "EF4444")
        cd.fill = fill_evt
        cd.alignment = Alignment(horizontal="left", vertical="center")
        cd.border = _THIN; cd.protection = _LOCK
        ws.row_dimensions[row].height = 18; row += 1

    _blank_row(ws, row, NC); row += 1

    # ── Section E: Top Users ─────────────────────────────────────────────
    _section_row(ws, row, "▌ E.  MOST ACTIVE USERS", NC); row += 1

    for ci, txt in enumerate(["User ID", "Events", "% Total", "Activity Bar"], 1):
        c = ws.cell(row, ci, txt)
        c.font = _F_COLHDR; c.fill = _FILL_COLHDR; c.alignment = _CENTER
        c.border = _THIN; c.protection = _LOCK
    ws.row_dimensions[row].height = 20; row += 1

    max_usr = top_users[0][1] if top_users else 1
    for rank, (user, cnt) in enumerate(top_users, 1):
        fill_u = _FILL_ZEBRA if rank % 2 else _FILL_WHITE

        ws.cell(row, 1, user).font = _F_MONO
        ws.cell(row, 1).fill = fill_u; ws.cell(row, 1).alignment = _LEFT
        ws.cell(row, 1).border = _THIN; ws.cell(row, 1).protection = _LOCK

        ws.cell(row, 2, f"{cnt:,}").font = _F_BOLD
        ws.cell(row, 2).fill = fill_u; ws.cell(row, 2).alignment = _CENTER
        ws.cell(row, 2).border = _THIN; ws.cell(row, 2).protection = _LOCK

        ws.cell(row, 3, _pct(cnt, total)).font = _F_MUTED
        ws.cell(row, 3).fill = fill_u; ws.cell(row, 3).alignment = _CENTER
        ws.cell(row, 3).border = _THIN; ws.cell(row, 3).protection = _LOCK

        bar_cell = ws.cell(row, 4, _bar(cnt, max_usr, 40))
        bar_cell.font = Font(name="Courier New", size=10, color="0EA5E9")
        bar_cell.fill = fill_u
        bar_cell.alignment = Alignment(horizontal="left", vertical="center")
        bar_cell.border = _THIN; bar_cell.protection = _LOCK
        ws.row_dimensions[row].height = 18; row += 1

    _blank_row(ws, row, NC); row += 1

    # ── Section F: Company Activity ──────────────────────────────────────
    if per_company:
        _section_row(ws, row, "▌ F.  COMPANY ACTIVITY COMPARISON", NC); row += 1

        for ci, txt in enumerate(["Company", "Total", "Errors/Warn", "Last 24h"], 1):
            c = ws.cell(row, ci, txt)
            c.font = _F_COLHDR; c.fill = _FILL_COLHDR; c.alignment = _CENTER
            c.border = _THIN; c.protection = _LOCK
        ws.row_dimensions[row].height = 20; row += 1

        companies_sorted = sorted(per_company.values(), key=lambda x: x.get("total", 0), reverse=True)
        for ri, co in enumerate(companies_sorted):
            fill_co = _FILL_ZEBRA if ri % 2 else _FILL_WHITE
            errs_co = co.get("errors_warnings", 0)

            ws.cell(row, 1, co.get("company_name", "?")).font = _F_BOLD
            ws.cell(row, 1).fill = fill_co; ws.cell(row, 1).alignment = _LEFT
            ws.cell(row, 1).border = _THIN; ws.cell(row, 1).protection = _LOCK

            ws.cell(row, 2, f"{co.get('total', 0):,}").font = _F_BOLD
            ws.cell(row, 2).fill = fill_co; ws.cell(row, 2).alignment = _CENTER
            ws.cell(row, 2).border = _THIN; ws.cell(row, 2).protection = _LOCK

            c3 = ws.cell(row, 3, f"{errs_co:,}")
            c3.font = Font(name="Calibri", bold=True, size=10, color="991B1B" if errs_co > 0 else "166534")
            c3.fill = fill_co; c3.alignment = _CENTER; c3.border = _THIN; c3.protection = _LOCK

            ws.cell(row, 4, f"{co.get('recent_24h', 0):,}").font = _F_DATA
            ws.cell(row, 4).fill = fill_co; ws.cell(row, 4).alignment = _CENTER
            ws.cell(row, 4).border = _THIN; ws.cell(row, 4).protection = _LOCK
            ws.row_dimensions[row].height = 20; row += 1

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    _lock(ws)


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 3 — Alerts & Anomalies
# ═══════════════════════════════════════════════════════════════════════

def _build_alerts_sheet(ws, logs: list, ts: str):
    alerts = [l for l in logs if (l.get("severity") or "info") in ("warning", "error", "critical")]
    nc = len(_S1_COLS)

    _header_block(
        ws,
        "ALERTS & ANOMALIES",
        f"Generated: {ts}  |  Warning / Error / Critical only  |  {len(alerts):,} records",
        nc,
    )

    if not alerts:
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=nc)
        c = ws.cell(4, 1, "No alerts or anomalies detected in this report period.")
        c.font = _F_SAFE; c.alignment = _CENTER; c.border = _THIN; c.protection = _LOCK
        ws.row_dimensions[4].height = 28
        _lock(ws)
        return

    hdr = 4
    _col_headers(ws, hdr, _S1_COLS)

    for idx, log in enumerate(sorted(alerts, key=lambda x: (x.get("severity") or ""), reverse=True)):
        r = hdr + 1 + idx
        sev = (log.get("severity") or "warning").lower()
        bg, fg, bar_c, _ = _SEV.get(sev, _SEV["warning"])
        fill = PatternFill("solid", fgColor=bg)

        left_border = Border(
            left=Side("medium", color=bar_c),
            right=Side("thin", color="E2E8F0"),
            top=Side("thin", color="F1F5F9"),
            bottom=Side("thin", color="F1F5F9"),
        )

        vals = [
            idx + 1, _fmt_ts(log.get("created_at")), log.get("company_name") or "",
            log.get("user_id") or "—", (log.get("category") or "system").upper(), sev.upper(),
            log.get("event_type") or "", log.get("title") or "", log.get("detail") or "",
            (f"{log.get('entity_type') or ''}: {log.get('entity_name') or ''}").strip(": ") or "—",
            log.get("ip_address") or "—",
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(r, ci, val)
            c.fill = fill; c.protection = _LOCK
            c.border = left_border if ci == 1 else _THIN
            c.font = Font(name="Calibri", bold=(sev == "critical"), size=10, color=fg)
            c.alignment = _WRAP if ci in (8, 9) else (Alignment(vertical="center") if ci in (2, 7, 11) else _LEFT)
        ws.row_dimensions[r].height = 18

    last = hdr + max(len(alerts), 1)
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(nc)}{last}"
    ws.freeze_panes = f"A{hdr + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    _lock(ws)


# ═══════════════════════════════════════════════════════════════════════
#  SHEET 4 — Company Report
# ═══════════════════════════════════════════════════════════════════════

_S4_COLS = [
    ("#",                  5),
    ("Company",           28),
    ("Total Events",      14),
    ("Info Events",       13),
    ("Errors & Warnings", 18),
    ("Last 24h",          12),
    ("Error Rate",        12),
    ("Status",            16),
]


def _build_company_sheet(ws, stats: dict, ts: str):
    per_company = stats.get("per_company", {})
    nc = len(_S4_COLS)

    _header_block(ws, "COMPANY PERFORMANCE REPORT", f"Generated: {ts}  |  {len(per_company)} companies  |  Sorted by total activity", nc)

    hdr = 4
    _col_headers(ws, hdr, _S4_COLS)

    companies = sorted(per_company.values(), key=lambda x: x.get("total", 0), reverse=True)

    for idx, co in enumerate(companies):
        r = hdr + 1 + idx
        total_co = co.get("total", 0)
        errs_co  = co.get("errors_warnings", 0)   # warning+error+critical combined
        info_cnt = total_co - errs_co
        recent   = co.get("recent_24h", 0)
        err_rate = f"{errs_co / total_co * 100:.1f}%" if total_co else "—"

        status = "CRITICAL" if errs_co > 10 else ("ATTENTION" if errs_co > 0 else "NORMAL")
        status_color = {"CRITICAL": "991B1B", "ATTENTION": "92400E", "NORMAL": "166534"}[status]
        status_bg    = {"CRITICAL": "FEE2E2", "ATTENTION": "FFFBEB", "NORMAL": "DCFCE7"}[status]

        fill = _FILL_ZEBRA if idx % 2 else _FILL_WHITE

        row_vals = [
            idx + 1,
            co.get("company_name", "?"),
            f"{total_co:,}",
            f"{info_cnt:,}",
            f"{errs_co:,}",
            f"{recent:,}",
            err_rate,
            status,
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(r, ci, val)
            c.protection = _LOCK; c.border = _THIN

            if ci == 8:  # Status
                c.fill = PatternFill("solid", fgColor=status_bg)
                c.font = Font(name="Calibri", bold=True, size=10, color=status_color)
                c.alignment = _CENTER
            elif ci == 3:  # Total
                c.fill = fill
                c.font = Font(name="Calibri", bold=True, size=11, color=_BRAND_DARK)
                c.alignment = _CENTER
            elif ci == 5:  # Errors & Warnings
                c.fill = fill
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="991B1B" if errs_co > 0 else "166534")
                c.alignment = _CENTER
            else:
                c.fill = fill; c.font = _F_DATA
                c.alignment = _CENTER if ci != 2 else _LEFT
        ws.row_dimensions[r].height = 22

    last = hdr + max(len(companies), 1)
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(nc)}{last}"
    ws.freeze_panes = f"A{hdr + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    _lock(ws)


# ═══════════════════════════════════════════════════════════════════════
#  Public API
# ═══════════════════════════════════════════════════════════════════════

def generate_platform_logs_excel(
    logs: list,
    stats: dict,
    report_title: Optional[str] = None,
) -> io.BytesIO:
    """
    Generate enterprise platform logs Excel workbook.

    Args:
        logs:   List of log record dicts (company_name, user_id, severity, etc.)
        stats:  Stats dict from /api/admin/platform-logs/stats
        report_title: Optional custom title override

    Returns:
        BytesIO containing the .xlsx workbook, ready for send_file().
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Activity Log"
    _build_activity_sheet(ws1, logs, ts, stats.get("total", len(logs)))

    ws2 = wb.create_sheet("System Intelligence")
    _build_intelligence_sheet(ws2, logs, stats, ts)

    ws3 = wb.create_sheet("Alerts & Anomalies")
    _build_alerts_sheet(ws3, logs, ts)

    ws4 = wb.create_sheet("Company Report")
    _build_company_sheet(ws4, stats, ts)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
